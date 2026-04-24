"""from openpyxl import Workbook
w=Workbook()
sheet=w.active
sheet.title="Users"

sheet.append(["ID","Name","Role"])
w.save("library.xlsx")

from openpyxl import load_workbook
w=load_workbook("library.xlsx")
sheet=w["Users"]
sheet.append([1,"vishal","student"])
sheet.append([2,"Admin1","Admin"])
w.save("library.xlsx")"""

"""page=int(input("enter page:"))
size=int(input("enter size:"))
skip=(page*size)-size
print("skipped file:",skip)"""

import mysql.connector
conn=mysql.connector.connect(host="localhost",user="root",password="Vishal@12",database="library")
cursor=conn.cursor()
cursor.execute("Alter table users add unique (name)")
conn.commit()
print("success")
conn.close()